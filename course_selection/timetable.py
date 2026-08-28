"""Import a normalized school timetable into a local, conflict-ready model."""

from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable


@dataclass(frozen=True)
class TimetableEntry:
    term: str
    course_code: str
    course_name: str
    weekday: int | None
    start_period: int | None
    end_period: int | None
    week_start: int | None
    week_end: int | None
    week_parity: str
    teaching_class: str = ""
    teacher: str = ""
    location: str = ""
    week_numbers: tuple[int, ...] = ()
    # ``unknown`` is deliberately distinct from an empty/free slot.  It is
    # set when a source explicitly reports a pending time or cannot be
    # parsed reliably, so conflict checking must ask the user to review it.
    conflict_status: str = "known"


HEADER_ALIASES = {
    "term": ("学年学期", "学期", "教学学期"),
    "course_code": ("课程代码", "课程编号", "课程号"),
    "course_name": ("课程名称", "课程名"),
    "weekday": ("星期", "上课星期"),
    "periods": ("节次", "上课节次", "上课时间"),
    "weeks": ("周次", "上课周次", "教学周"),
    "week_parity": ("单双周", "单周/双周"),
    "location": ("上课地点", "教室", "地点"),
}


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _header_map(row: Iterable[object]) -> dict[str, int]:
    cells = [_clean(value) for value in row]
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for index, cell in enumerate(cells):
            if cell in aliases:
                result[field] = index
                break
    return result


def _parse_int_range(value: str, label: str) -> tuple[int, int]:
    numbers = [int(item) for item in re.findall(r"\d+", value)]
    if not numbers:
        raise ValueError(f"无法解析{label}：{value}")
    if len(numbers) == 1:
        return numbers[0], numbers[0]
    return min(numbers[0], numbers[1]), max(numbers[0], numbers[1])


_UNKNOWN_TIME = re.compile(r"待定|未定|待安排|未知|时间未定|tbd|pending", re.I)


def _is_unknown_time(value: str) -> bool:
    return bool(_UNKNOWN_TIME.search(value.strip()))


def _parse_week_numbers(value: str) -> tuple[int, ...]:
    normalized = value.replace("，", ",").replace("、", ",")
    numbers: set[int] = set()
    for match in re.finditer(r"(\d+)\s*(?:[-~至])\s*(\d+)|(\d+)", normalized):
        if match.group(1):
            start, end = int(match.group(1)), int(match.group(2))
            numbers.update(range(min(start, end), max(start, end) + 1))
        else:
            numbers.add(int(match.group(3)))
    return tuple(sorted(numbers))


def _parse_weekday(value: str) -> int:
    normalized = value.replace("星期", "").replace("周", "")
    chinese = {"一": "1", "二": "2", "三": "3", "四": "4", "五": "5", "六": "6", "日": "7", "天": "7"}
    normalized = chinese.get(normalized, normalized)
    match = re.search(r"([1-7])", normalized)
    if not match:
        raise ValueError(f"无法解析星期：{value}")
    return int(match.group(1))


def _parse_parity(value: str) -> str:
    if "单" in value:
        return "odd"
    if "双" in value:
        return "even"
    return "all"


def normalize_timetable_term(value: str) -> str:
    match = re.search(r"(20\d{2})\s*(春季|夏季|秋季|冬季)学期", value)
    if match:
        return f"{match.group(1)}年{match.group(2)}学期"
    return value.strip()


def _value(row: list[object], headers: dict[str, int], field: str) -> str:
    index = headers.get(field)
    return _clean(row[index]) if index is not None and index < len(row) else ""


def _rows_from_workbook(path: Path) -> list[list[object]]:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return [list(row) for row in workbook.active.iter_rows(values_only=True)]
        finally:
            workbook.close()
    if path.suffix.lower() == ".xls":
        import xlrd

        sheet = xlrd.open_workbook(path, on_demand=True).sheet_by_index(0)
        return [sheet.row_values(index) for index in range(sheet.nrows)]
    raise ValueError("课表必须是 .xls 或 .xlsx 文件")


def _is_grid_timetable(rows: list[list[object]]) -> bool:
    if len(rows) <= 1:
        return False
    weekday_headers = sum(
        _clean(value).startswith("星期") for value in rows[1]
    )
    return weekday_headers >= 2


def _grid_entry_text(value: str) -> tuple[str, str, str, str]:
    lines = [line.strip() for line in re.split(r"</?br>|\n", value) if line.strip()]
    if not lines:
        raise ValueError("课表单元格没有课程内容")
    first = lines[0]
    combined = " ".join(lines)
    if "◇" in first:
        parts = [part.strip() for part in first.split("◇")]
        course_name = parts[0]
        week_text = next((part for part in parts if "周" in part or re.search(r"\d", part)), "")
        teacher = re.sub(r"\[.*", "", week_text).strip() if "[" in week_text else ""
        location = parts[-1] if len(parts) >= 3 and parts[-1] != week_text else ""
    else:
        course_name = first
        schedule_text = " ".join(lines[1:])
        # 收集所有 [周次] 区间,支持多教师/多区间(如 [1-5，7-9]周，[16]周)。
        week_parts = re.findall(r"\[([^]]+)\]", schedule_text)
        week_text = "，".join(part for part in week_parts)
        teacher_parts = re.findall(
            r"(?:^|周[，,、；;\s]*)([^][，,、；;]+?)\s*\[", schedule_text
        )
        teacher = "，".join(part.strip() for part in teacher_parts if part.strip())
        location = ""
        if week_parts:
            if len(lines) >= 3 and not re.search(r"\[", lines[-1]):
                # 地点独占最后一行: "课程名<br>教师[周次]...<br>地点"
                location = lines[-1].replace("周", "").strip()
            else:
                # 地点紧跟最后一个周次: "课程名<br>教师[周次]周地点"
                matches = list(re.finditer(r"\[[^\]]*\]\s*周", schedule_text))
                if matches:
                    location = schedule_text[matches[-1].end():].strip()
    if not week_text and _is_unknown_time(value):
        week_text = "待定"
    if not week_text:
        raise ValueError(f"课程缺少教学周：{value}")
    return course_name, week_text, teacher, location


def import_grid_timetable_rows(
    rows: list[list[object]], expected_term: str | None = None
) -> tuple[TimetableEntry, ...]:
    """Adapt normalized grid rows from HTML or workbook sources."""
    term = normalize_timetable_term(_clean(rows[0][0]) if rows and rows[0] else "")
    if expected_term and term != expected_term:
        raise ValueError(f"课表学期不匹配：发现 {term!r}，期望 {expected_term!r}")
    weekdays = {
        column: _parse_weekday(_clean(value))
        for column, value in enumerate(rows[1])
        if _clean(value).startswith("星期")
    }
    entries: list[TimetableEntry] = []
    for row in rows[2:]:
        if len(row) < 2 or not _clean(row[1]):
            continue
        period_text = _clean(row[1])
        unknown_period = _is_unknown_time(period_text)
        if unknown_period:
            period_start = period_end = None
        else:
            period_start, period_end = _parse_int_range(period_text, "节次")
        for column, weekday in weekdays.items():
            if column >= len(row) or not _clean(row[column]):
                continue
            for raw_line in re.split(r"</?br>", _clean(row[column])):
                if not raw_line.strip():
                    continue
                course_name, week_text, teacher, location = _grid_entry_text(raw_line)
                week_numbers = _parse_week_numbers(week_text)
                unknown_time = unknown_period or _is_unknown_time(week_text) or not week_numbers
                if unknown_time:
                    week_numbers = ()
                entries.append(
                    TimetableEntry(
                        term=term,
                        course_code="",
                        course_name=course_name,
                        weekday=weekday,
                        start_period=period_start,
                        end_period=period_end,
                        week_start=min(week_numbers) if week_numbers else None,
                        week_end=max(week_numbers) if week_numbers else None,
                        week_parity=_parse_parity(week_text),
                        teacher=teacher,
                        location=location,
                        week_numbers=week_numbers,
                        conflict_status="unknown" if unknown_time else "known",
                    )
                )
    if not entries:
        raise ValueError("课表没有可用课程记录")
    return tuple(entries)


def import_timetable(path: Path, *, expected_term: str | None = None) -> tuple[TimetableEntry, ...]:
    rows = _rows_from_workbook(path)
    if _is_grid_timetable(rows):
        return import_grid_timetable_rows(rows, expected_term)
    header_index = next(
        (index for index, row in enumerate(rows) if "course_name" in _header_map(row)),
        None,
    )
    if header_index is None:
        raise ValueError("未找到包含课程名称的课表表头")
    headers = _header_map(rows[header_index])
    required = {"term", "course_name", "weekday", "periods", "weeks"}
    missing = required - headers.keys()
    if missing:
        raise ValueError(f"课表缺少必要列：{', '.join(sorted(missing))}")

    entries: list[TimetableEntry] = []
    for row in rows[header_index + 1 :]:
        if not any(_clean(value) for value in row):
            continue
        term = normalize_timetable_term(_value(row, headers, "term"))
        course_name = _value(row, headers, "course_name")
        if not term and not course_name:
            continue
        if expected_term and term != expected_term:
            raise ValueError(f"课表学期不匹配：发现 {term!r}，期望 {expected_term!r}")
        periods_text = _value(row, headers, "periods")
        weeks_text = _value(row, headers, "weeks")
        unknown_time = _is_unknown_time(periods_text) or _is_unknown_time(weeks_text)
        if unknown_time:
            period_start = period_end = week_start = week_end = None
            week_numbers = ()
        else:
            period_start, period_end = _parse_int_range(periods_text, "节次")
            week_start, week_end = _parse_int_range(weeks_text, "周次")
            week_numbers = _parse_week_numbers(weeks_text)
        entries.append(
            TimetableEntry(
                term=term,
                course_code=_value(row, headers, "course_code"),
                course_name=course_name,
                weekday=_parse_weekday(_value(row, headers, "weekday")),
                start_period=period_start,
                end_period=period_end,
                week_start=week_start,
                week_end=week_end,
                week_parity=_parse_parity(_value(row, headers, "week_parity")),
                location=_value(row, headers, "location"),
                week_numbers=week_numbers,
                conflict_status="unknown" if unknown_time else "known",
            )
        )
    if not entries:
        raise ValueError("课表没有可用课程记录")
    return tuple(entries)


def entries_to_dict(entries: Iterable[TimetableEntry]) -> list[dict[str, object]]:
    return [asdict(entry) for entry in entries]


def timetable_snapshot_payload(
    entries: Iterable[TimetableEntry], *, source_name: str, source_kind: str = "user-imported",
    source_at: str | None = None,
) -> dict[str, object]:
    """Build the persisted, explicitly sourced timetable representation.

    Keeping this construction beside the importer prevents a file import from
    being mistaken for a live academic read.  Unknown time rows remain in the
    snapshot and carry ``conflict_status=unknown``.
    """
    materialized = tuple(entries)
    if not materialized:
        raise ValueError("timetable has no usable course records")
    payload: dict[str, object] = {
        "status": "complete",
        "source_kind": source_kind,
        "source_name": source_name,
        "entries": entries_to_dict(materialized),
    }
    if source_at:
        payload["source_at"] = source_at
    return payload
